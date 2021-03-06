"""Contain ZIP import controller."""
import ckan.lib.base as base
import ckan.lib.helpers as h
from ckan.lib.munge import munge_title_to_name
from ckan.common import request, c, _
import ckan.model as model
from ckan.logic import (ValidationError, NotAuthorized,
                        NotFound, check_access)
from ckan.logic.validators import package_id_or_name_exists
import ckan.lib.navl.dictization_functions as df
import ckan.plugins.toolkit as tk

import zipfile
from openpyxl import load_workbook
import io

from ckanext.excelimport import (
    AVAILABLE_MD_FILES,
    METADATA_SHEET
)
from ckanext.excelimport.helpers import get_helpers

abort = base.abort
Invalid = df.Invalid


class ExcelImportController(base.BaseController):
    """Controller that provides seed datasets import from ZIPed xlsx files."""

    def run_create(self, context, data_dict, resources_sheet, archive):
        """Dataset creating proccess."""
        data_dict['name'] = munge_title_to_name(
            data_dict['title']
        )
        try:
            package_id_or_name_exists(data_dict['name'], context)
        except Invalid:
            pass
        else:
            counter = 0

            while True:
                name = '{0}-{1}'.format(data_dict['name'], counter)
                try:
                    package_id_or_name_exists(name, context)
                except Invalid:
                    data_dict['name'] = name
                    break
                counter += 1

        result = self.create_dataset(
            context,
            data_dict,
            resources_sheet,
            archive
        )

        if result:
            h.flash_success('Dataset was created!')

    def import_from_zip(self):
        """Method that renders the form and receive submition of the form."""
        context = {
            'model': model,
            'session': model.Session,
            'user': c.user or c.author,
            'for_view': True,
            'auth_user_obj': c.userobj
        }

        try:
            check_access('package_create', context)
        except NotAuthorized:
            abort(401, _('Unauthorized to create package'))
        except NotFound:
            abort(404, _('Dataset not found'))

        if request.method == 'POST':

            try:
                zip_file = request.params.get('dataset_zip').filename
                get_helpers.get('validate_file_ext')(zip_file)
            except ValidationError, e:
                h.flash_error(e.error_dict['message'])
            except AttributeError, e:
                h.flash_error('Upload field is empty')
            else:
                archive = zipfile.ZipFile(
                    request.params.get('dataset_zip').file,
                    'r'
                )
                list_files = archive.namelist()
                md_file = [i for i in AVAILABLE_MD_FILES if i in list_files]

                if md_file:
                    metadata = archive.read(md_file[0])
                    metadata_xlsx = load_workbook(
                        filename=io.BytesIO(metadata),
                        data_only=True
                    )

                    # Get metadata sheet
                    try:
                        metadata_sheet = metadata_xlsx.get_sheet_by_name(
                            METADATA_SHEET[md_file[0]]
                        )
                    except KeyError, e:
                        h.flash_error(e)
                    else:
                        #  Get resources sheet
                        resources_sheet = metadata_xlsx.get_sheet_by_name(
                            'Resources'
                        )
                        data_dict = {}
                        rows = metadata_sheet.iter_rows(row_offset=1)
                        try:
                            get_helpers.get('prepare_data_dict')(data_dict, rows)
                        except KeyError, e:
                            h.flash_error('Not mapped field: {0}'.format(e))

                        if not data_dict.get('id', False):
                            self.run_create(
                                context,
                                data_dict,
                                resources_sheet,
                                archive
                            )
                        else:
                            pkg = tk.get_action('package_show')(
                                None,
                                {'id': data_dict['id']}
                            )
                            if pkg:
                                h.flash_error('Dataset with id {0} exists'.format(data_dict['id']))
                            else:
                                del data_dict['id']
                                self.run_create(
                                    context,
                                    data_dict,
                                    resources_sheet,
                                    archive
                                )

                else:
                    h.flash_error(
                        'ZIP must contain 1 of 2 files: {0} or {1}'.format(
                            AVAILABLE_MD_FILES[0],
                            AVAILABLE_MD_FILES[1]
                        )
                    )

        return base.render('snippets/import-from-zip.html')

    def create_dataset(self, context, data_dict, res_sheet, archive):
        """Create dataset with resources (if it has any)."""
        try:
            ds = tk.get_action('package_create')(context, data_dict)
            if res_sheet:
                self.create_resource(context, ds, res_sheet, archive)

            return ds
        except ValidationError, e:
            h.flash_error(e.error_dict)

    def create_resource(self, context, data_dict, sheet, archive):
        """Create resource with file source or url source."""
        rows = sheet.iter_rows(row_offset=1)
        list_files = archive.namelist()

        for row in rows:
            resource_from = row[1].value
            resource_format = row[2].value
            resource_title = row[3].value
            resource_desc = row[5].value
            if resource_from:
                if not resource_from.startswith('http'):
                    if resource_from in list_files:
                        read_data = archive.read(resource_from)
                        parse_data = io.BytesIO(read_data)
                        fs = get_helpers.get('prepare_file')(
                            read_data,
                            parse_data,
                            resource_from
                        )

                        tk.get_action('resource_create')(context, {
                            'package_id': data_dict['id'],
                            # url must be provided, even for uploads
                            'url': resource_from,
                            'format': resource_format,
                            'name': resource_title,
                            'description': resource_desc,
                            'url_type': 'upload',
                            'upload': fs
                        })

                elif resource_from:
                    tk.get_action('resource_create')(context, {
                        'package_id': data_dict['id'],
                        'url': resource_from,
                        'name': resource_title,
                        'format': resource_format,
                        'description': resource_desc,
                    })
