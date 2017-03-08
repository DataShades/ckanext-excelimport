"""Contain ZIP import controller."""
import ckan.lib.base as base
import ckan.lib.helpers as h
from ckan.common import request, c, _
import ckan.model as model
from ckan.logic import (ValidationError, NotAuthorized,
                        NotFound, check_access)
import ckan.plugins.toolkit as tk

import zipfile
from openpyxl import load_workbook
import io

from ckanext.excelimport import (
    AVAILABLE_MD_FILES,
    METADATA_SHEET
)
from ckanext.excelimport.helpers import get_helpers

abort = base.abort


class ExcelUpdateController(base.BaseController):
    """Controller that provides seed datasets import from ZIPed xlsx files."""

    def update_from_zip(self, id):
        """Method that renders the form and receive submition of the form."""
        context = {
            'model': model,
            'session': model.Session,
            'user': c.user or c.author,
            'for_view': True,
            'auth_user_obj': c.userobj
        }

        try:
            check_access('package_update', context, {'id': id})
            c.pkg_dict = tk.get_action('package_show')(context, {'id': id})
        except NotAuthorized:
            abort(401, _('Unauthorized to read package %s') % '')
        except NotFound:
            abort(404, _('Dataset not found'))

        c.pkg = context.get("package")

        if request.method == 'POST':
            try:
                zip_file = request.params.get('dataset_zip').filename
                get_helpers.get('validate_file_ext')(zip_file)
            except ValidationError, e:
                h.flash_error(e.error_dict['message'])
            except AttributeError, e:
                h.flash_error('Upload field is empty')
            else:
                archive = zipfile.ZipFile(
                    request.params.get('dataset_zip').file,
                    'r'
                )
                list_files = archive.namelist()
                md_file = [i for i in AVAILABLE_MD_FILES if i in list_files]

                if md_file:
                    metadata = archive.read(md_file[0])
                    metadata_xlsx = load_workbook(
                        filename=io.BytesIO(metadata),
                        data_only=True
                    )

                    # Get metadata sheet
                    try:
                        metadata_sheet = metadata_xlsx.get_sheet_by_name(
                            METADATA_SHEET[md_file[0]]
                        )
                    except KeyError, e:
                        h.flash_error(e)
                    else:
                        #  Get resources sheet
                        resources_sheet = metadata_xlsx.get_sheet_by_name(
                            'Resources'
                        )
                        data_dict = {}
                        rows = metadata_sheet.iter_rows(row_offset=1)
                        try:
                            get_helpers.get('prepare_data_dict')(data_dict, rows)
                        except KeyError, e:
                            h.flash_error('Not mapped field: {0}'.format(e))

                        if not data_dict.get('id', False):
                            h.flash_error('Metadata sheet must contain package id field.')
                        if data_dict['id'] not in [c.pkg_dict['id'], c.pkg_dict['name']]:
                            error_msg = (
                                '<p>Metadata sheet id field is incorrect.</p>'
                                '<i>Sheet id can be dataset\'s name or id.</i>'
                                '<br/>sheet id: {0}<br/>'
                                'datatest id: {1}<br/>'
                                'dataset name: {2}'
                            ).format(
                                data_dict['id'],
                                c.pkg_dict['id'],
                                c.pkg_dict['name']
                            )
                            h.flash_error(error_msg, True)
                        else:
                            result = self.update_dataset(
                                context,
                                resources_sheet,
                                archive,
                                data_dict,
                                c.pkg_dict
                            )
                            if result:
                                h.flash_success('Dataset was updated!')

                else:
                    h.flash_error(
                        'ZIP must contain 1 of 2 files: {0} or {1}'.format(
                            AVAILABLE_MD_FILES[0],
                            AVAILABLE_MD_FILES[1]
                        )
                    )

        return base.render('snippets/update-from-zip.html')

    def update_dataset(self, context, res_sheet, archive, data_dict, pkg_dict):
        """Update dataset that already exists."""
        try:
            pkg_dict.update(data_dict)
            del pkg_dict['resources']
            ds = tk.get_action('package_update')(context, pkg_dict)
            if res_sheet:
                self.create_resource(context, ds, res_sheet, archive)

            return ds
        except ValidationError, e:
            h.flash_error(e.error_dict)

    def create_resource(self, context, data_dict, sheet, archive):
        """Create resource with file source or url source."""
        rows = sheet.iter_rows(row_offset=1)
        list_files = archive.namelist()

        for row in rows:
            resource_from = row[1].value
            resource_format = row[2].value
            resource_title = row[3].value
            resource_desc = row[5].value
            if resource_from:
                if not resource_from.startswith('http'):
                    if resource_from in list_files:
                        read_data = archive.read(resource_from)
                        parse_data = io.BytesIO(read_data)
                        fs = get_helpers.get('prepare_file')(
                            read_data,
                            parse_data,
                            resource_from
                        )

                        tk.get_action('resource_create')(context, {
                            'package_id': data_dict['id'],
                            # url must be provided, even for uploads
                            'url': resource_from,
                            'name': resource_title,
                            'description': resource_desc,
                            'url_type': 'upload',
                            'upload': fs
                        })

                elif resource_from:
                    tk.get_action('resource_create')(context, {
                        'package_id': data_dict['id'],
                        'url': resource_from,
                        'name': resource_title,
                        'format': resource_format,
                        'description': resource_desc,
                    })
