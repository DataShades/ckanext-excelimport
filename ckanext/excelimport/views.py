import zipfile
import io

from openpyxl import load_workbook
from flask import Blueprint

import ckan.plugins.toolkit as tk
import ckan.lib.base as base
import ckan.lib.helpers as h
import ckan.model as model
import ckan.lib.navl.dictization_functions as df

from ckan.common import c, request, _, g
from ckan.lib.munge import munge_title_to_name
from ckan.logic import (ValidationError, NotAuthorized,
                        NotFound, check_access)
from ckan.logic.validators import package_id_or_name_exists


from ckanext.excelimport import AVAILABLE_MD_FILES, METADATA_SHEET
from ckanext.excelimport.helpers import get_helpers

abort = base.abort
Invalid = df.Invalid

excelimport = Blueprint('excelimport', __name__)


def _run_create(context, data_dict, resources_sheet, archive):
    """Dataset creating proccess."""
    data_dict['name'] = munge_title_to_name(
        data_dict['title']
    )
    try:
        package_id_or_name_exists(data_dict['name'], context)
    except Invalid:
        pass
    else:
        counter = 0

        while True:
            name = '{0}-{1}'.format(data_dict['name'], counter)
            try:
                package_id_or_name_exists(name, context)
            except Invalid:
                data_dict['name'] = name
                break
            counter += 1

    data_dict['owner_org'] = data_dict['owner_org'].lower()

    result = _create_dataset(
        context,
        data_dict,
        resources_sheet,
        archive
    )

    if result:
        h.flash_success('Dataset was created!')


def import_from_zip():
    """Method that renders the form and receive submition of the form."""
    context = {
        'model': model,
        'session': model.Session,
        'user': c.user or c.author,
        'for_view': True,
        'auth_user_obj': c.userobj
    }

    try:
        check_access('package_create', context)
    except NotAuthorized:
        abort(401, _('Unauthorized to create package'))
    except NotFound:
        abort(404, _('Dataset not found'))

    if request.method == 'POST':

        try:
            zip_file = request.files.get('dataset_zip').filename
            get_helpers.get('validate_file_ext')(zip_file)
        except ValidationError as e:
            h.flash_error(e.error_dict['message'])
        except AttributeError as e:
            h.flash_error('Upload field is empty')
        else:
            archive = zipfile.ZipFile(
                request.files.get('dataset_zip')._file,
                'r'
            )
            list_files = archive.namelist()
            md_file = [i for i in AVAILABLE_MD_FILES if i in list_files]

            if md_file:
                metadata = archive.read(md_file[0])
                metadata_xlsx = load_workbook(
                    filename=io.BytesIO(metadata),
                    data_only=True
                )

                # Get metadata sheet
                try:
                    metadata_sheet = metadata_xlsx[METADATA_SHEET[md_file[0]]]
                except KeyError as e:
                    h.flash_error(e)
                else:
                    #  Get resources sheet
                    resources_sheet = metadata_xlsx['Resources']
                    data_dict = {}
                    rows = metadata_sheet.iter_rows(row_offset=1)
                    try:
                        get_helpers.get('prepare_data_dict')(data_dict, rows)
                    except KeyError as e:
                        h.flash_error('Not mapped field: {0}'.format(e))

                    if not data_dict.get('id', False):
                        _run_create(
                            context,
                            data_dict,
                            resources_sheet,
                            archive
                        )
                    else:
                        pkg = tk.get_action('package_show')(
                            None,
                            {'id': data_dict['id']}
                        )
                        if pkg:
                            h.flash_error('Dataset with id {0} exists'.format(data_dict['id']))
                        else:
                            del data_dict['id']
                            _run_create(
                                context,
                                data_dict,
                                resources_sheet,
                                archive
                            )

            else:
                h.flash_error(
                    'There are no necessary files in the zip: {}'.format(
                        ', '.join(AVAILABLE_MD_FILES)
                    )
                )
    data = {"user_dict": g.userobj}
    return base.render('snippets/import-from-zip.html', extra_vars=data)


def _create_dataset(context, data_dict, res_sheet, archive):
    """Create dataset with resources (if it has any)."""
    try:
        ds = tk.get_action('package_create')(context, data_dict)
        if res_sheet:
            _create_resource(context, ds, res_sheet, archive)

        return ds
    except ValidationError as e:
        h.flash_error(e.error_dict)


def _create_resource(context, data_dict, sheet, archive):
    """Create resource with file source or url source."""
    rows = sheet.iter_rows(row_offset=1)
    list_files = archive.namelist()

    for row in rows:
        resource_from = row[1].value
        resource_format = row[2].value
        resource_title = row[3].value
        resource_desc = row[5].value
        if resource_from:
            if not resource_from.startswith('http'):
                if resource_from in list_files:
                    read_data = archive.read(resource_from)
                    parse_data = io.BytesIO(read_data)
                    fs = get_helpers.get('prepare_file')(
                        read_data,
                        parse_data,
                        resource_from
                    )

                    tk.get_action('resource_create')(context, {
                        'package_id': data_dict['id'],
                        # url must be provided, even for uploads
                        'url': resource_from,
                        'format': resource_format,
                        'name': resource_title,
                        'description': resource_desc,
                        'url_type': 'upload',
                        'upload': fs
                    })

            elif resource_from:
                tk.get_action('resource_create')(context, {
                    'package_id': data_dict['id'],
                    'url': resource_from,
                    'name': resource_title,
                    'format': resource_format,
                    'description': resource_desc,
                })


def update_from_zip(id):
    """Method that renders the form and receive submition of the form."""
    context = {
        'model': model,
        'session': model.Session,
        'user': c.user or c.author,
        'for_view': True,
        'auth_user_obj': c.userobj
    }

    try:
        check_access('package_update', context, {'id': id})
        c.pkg_dict = tk.get_action('package_show')(context, {'id': id})
    except NotAuthorized:
        abort(401, _('Unauthorized to read package %s') % '')
    except NotFound:
        abort(404, _('Dataset not found'))

    c.pkg = context.get("package")

    if request.method == 'POST':
        try:
            zip_file = request.files.get('dataset_zip').filename
            get_helpers.get('validate_file_ext')(zip_file)
        except ValidationError as e:
            h.flash_error(e.error_dict['message'])
        except AttributeError as e:
            h.flash_error('Upload field is empty')
        else:
            archive = zipfile.ZipFile(
                request.files.get('dataset_zip')._file,
                'r'
            )
            list_files = archive.namelist()
            md_file = [i for i in AVAILABLE_MD_FILES if i in list_files]

            if md_file:
                metadata = archive.read(md_file[0])
                metadata_xlsx = load_workbook(
                    filename=io.BytesIO(metadata),
                    data_only=True
                )

                # Get metadata sheet
                try:
                    metadata_sheet = metadata_xlsx[METADATA_SHEET[md_file[0]]]
                except KeyError as e:
                    h.flash_error(e)
                else:
                    #  Get resources sheet
                    resources_sheet = metadata_xlsx['Resources']
                    data_dict = {}
                    rows = metadata_sheet.iter_rows(row_offset=1)
                    try:
                        get_helpers.get('prepare_data_dict')(data_dict, rows)
                    except KeyError as e:
                        h.flash_error('Not mapped field: {0}'.format(e))

                    if not data_dict.get('id', False):
                        h.flash_error('Metadata sheet must contain package id field.')
                    if data_dict['id'] not in [c.pkg_dict['id'], c.pkg_dict['name']]:
                        error_msg = (
                            '<p>Metadata sheet id field is incorrect.</p>'
                            '<i>Sheet id can be dataset\'s name or id.</i>'
                            '<br/>sheet id: {0}<br/>'
                            'datatest id: {1}<br/>'
                            'dataset name: {2}'
                        ).format(
                            data_dict['id'],
                            c.pkg_dict['id'],
                            c.pkg_dict['name']
                        )
                        h.flash_error(error_msg, True)
                    else:
                        result = _update_dataset(
                            context,
                            resources_sheet,
                            archive,
                            data_dict,
                            c.pkg_dict
                        )
                        if result:
                            h.flash_success('Dataset was updated!')

            else:
                h.flash_error(
                    'There are no necessary files in the zip: {}'.format(
                        ', '.join(AVAILABLE_MD_FILES)
                    )
                )
    data = {"user_dict": g.userobj, 'pkg_dict': c.pkg_dict}
    return base.render('snippets/update-from-zip.html', extra_vars=data)


def _update_dataset(context, res_sheet, archive, data_dict, pkg_dict):
        """Update dataset that already exists."""
        try:
            pkg_dict.update(data_dict)
            del pkg_dict['resources']
            ds = tk.get_action('package_update')(context, pkg_dict)
            if res_sheet:
                _create_resource(context, ds, res_sheet, archive)

            return ds
        except ValidationError as e:
            h.flash_error(e.error_dict)


excelimport.add_url_rule(u'/import-from-zip', view_func=import_from_zip, methods=(u'GET', u'POST'))
excelimport.add_url_rule(u'/update-from-zip/<id>/', view_func=update_from_zip, methods=(u'GET', u'POST'))


def get_blueprints():
    return [excelimport]
